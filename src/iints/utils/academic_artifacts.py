from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping, Sequence


@dataclass(frozen=True)
class AcademicTheme:
    navy: tuple[int, int, int] = (31, 47, 61)
    blue: tuple[int, int, int] = (38, 89, 128)
    teal: tuple[int, int, int] = (42, 157, 143)
    red: tuple[int, int, int] = (196, 64, 64)
    amber: tuple[int, int, int] = (224, 151, 56)
    paper: tuple[int, int, int] = (255, 255, 252)
    soft_blue: tuple[int, int, int] = (232, 241, 247)
    soft_gray: tuple[int, int, int] = (246, 247, 248)
    text: tuple[int, int, int] = (28, 35, 42)
    muted: tuple[int, int, int] = (96, 108, 119)
    rule: tuple[int, int, int] = (183, 192, 201)


ACADEMIC_THEME = AcademicTheme()


def _rgb_hex(rgb: tuple[int, int, int]) -> str:
    return "".join(f"{part:02X}" for part in rgb)


def _cell_next(pdf: Any, width: float, height: float, text: str, **kwargs: Any) -> None:
    try:
        from fpdf.enums import XPos, YPos

        pdf.cell(width, height, text, new_x=XPos.LMARGIN, new_y=YPos.NEXT, **kwargs)
    except Exception:
        pdf.cell(width, height, text, ln=1, **kwargs)


def _cell_right(pdf: Any, width: float, height: float, text: str, **kwargs: Any) -> None:
    try:
        from fpdf.enums import XPos, YPos

        pdf.cell(width, height, text, new_x=XPos.RIGHT, new_y=YPos.TOP, **kwargs)
    except Exception:
        pdf.cell(width, height, text, **kwargs)


def setup_academic_pdf(pdf: Any, *, title: str = "IINTS-AF Research Report") -> None:
    """Apply document metadata, margins, and base colors to an FPDF object."""
    try:
        pdf.set_title(title)
        pdf.set_author("IINTS-AF SDK")
        pdf.set_creator("IINTS-AF SDK")
        pdf.set_subject("Pre-clinical research simulation report")
    except Exception:
        pass
    pdf.set_margins(14, 14, 14)
    pdf.set_auto_page_break(auto=True, margin=16)
    pdf.set_text_color(*ACADEMIC_THEME.text)
    pdf.set_draw_color(*ACADEMIC_THEME.rule)


def add_academic_header(
    pdf: Any,
    title: str,
    *,
    subtitle: str = "Pre-clinical research report - not for treatment decisions",
    metadata: Mapping[str, Any] | None = None,
) -> None:
    """Render an academic title block with compact metadata."""
    theme = ACADEMIC_THEME
    pdf.set_x(pdf.l_margin)
    pdf.set_fill_color(*theme.paper)
    pdf.set_text_color(*theme.navy)
    pdf.set_font("Helvetica", "B", 16)
    pdf.multi_cell(0, 7.5, title)
    pdf.set_x(pdf.l_margin)
    pdf.set_font("Helvetica", "", 8.5)
    pdf.set_text_color(*theme.muted)
    pdf.multi_cell(0, 4.5, subtitle)

    metadata = dict(metadata or {})
    metadata.setdefault("Generated UTC", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M"))
    if metadata:
        pdf.ln(1)
        pdf.set_font("Helvetica", "", 7.5)
        pdf.set_text_color(*theme.muted)
        for idx, (key, value) in enumerate(metadata.items()):
            label = f"{key}: {value}"
            _cell_right(pdf, 62 if idx < 2 else 0, 4, str(label)[:55])
            if idx == 1:
                pdf.ln(4)
        if len(metadata) % 2:
            pdf.ln(4)

    y = pdf.get_y() + 2
    pdf.set_draw_color(*theme.blue)
    pdf.set_line_width(0.5)
    pdf.line(pdf.l_margin, y, pdf.w - pdf.r_margin, y)
    pdf.set_line_width(0.2)
    pdf.ln(6)
    pdf.set_text_color(*theme.text)


def add_academic_section(pdf: Any, title: str) -> None:
    """Render a clear section divider for FPDF reports."""
    theme = ACADEMIC_THEME
    pdf.ln(2)
    pdf.set_fill_color(*theme.navy)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 9)
    _cell_next(pdf, 0, 6, title.upper(), fill=True)
    pdf.set_text_color(*theme.text)
    pdf.ln(1)


def add_metric_cards(pdf: Any, cards: Sequence[tuple[str, str]], *, columns: int = 3) -> None:
    """Render compact metric cards that remain readable on printed PDFs."""
    theme = ACADEMIC_THEME
    columns = max(1, min(columns, 4))
    usable_width = pdf.w - pdf.l_margin - pdf.r_margin
    gutter = 4
    card_w = (usable_width - gutter * (columns - 1)) / columns
    card_h = 17
    start_x = pdf.l_margin
    start_y = pdf.get_y() + 1
    pdf.set_draw_color(*theme.rule)
    for idx, (label, value) in enumerate(cards):
        row = idx // columns
        col = idx % columns
        x = start_x + col * (card_w + gutter)
        y = start_y + row * (card_h + 4)
        pdf.set_fill_color(*theme.soft_blue)
        pdf.rect(x, y, card_w, card_h, style="FD")
        pdf.set_xy(x + 2.5, y + 2.3)
        pdf.set_text_color(*theme.muted)
        pdf.set_font("Helvetica", "", 7.2)
        pdf.cell(card_w - 5, 4, label[:34])
        pdf.set_xy(x + 2.5, y + 7.4)
        pdf.set_text_color(*theme.navy)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(card_w - 5, 7, value[:24])
    rows = (len(cards) + columns - 1) // columns
    pdf.set_y(start_y + rows * (card_h + 4))
    pdf.set_text_color(*theme.text)


def add_key_value_table(pdf: Any, rows: Sequence[tuple[str, str]], *, key_width: float = 62) -> None:
    theme = ACADEMIC_THEME
    value_width = pdf.w - pdf.l_margin - pdf.r_margin - key_width
    pdf.set_font("Helvetica", "", 8.5)
    for idx, (key, value) in enumerate(rows):
        fill = idx % 2 == 0
        if fill:
            pdf.set_fill_color(*theme.soft_gray)
        pdf.set_text_color(*theme.navy)
        pdf.set_font("Helvetica", "B", 8.5)
        pdf.cell(key_width, 5.5, key, fill=fill)
        pdf.set_text_color(*theme.text)
        pdf.set_font("Helvetica", "", 8.5)
        _cell_next(pdf, value_width, 5.5, value, fill=fill)


def add_academic_footer(pdf: Any, *, note: str | None = None) -> None:
    theme = ACADEMIC_THEME
    y = pdf.h - 12
    pdf.set_y(y)
    pdf.set_draw_color(*theme.rule)
    pdf.line(pdf.l_margin, y - 1, pdf.w - pdf.r_margin, y - 1)
    pdf.set_font("Helvetica", "I", 7)
    pdf.set_text_color(*theme.muted)
    pdf.multi_cell(
        0,
        3.3,
        note
        or "IINTS-AF SDK research artifact. Pre-clinical simulation output; not a medical-device report or treatment recommendation.",
    )
    pdf.set_text_color(*theme.text)


def style_excel_workbook(path: str | Path, *, title: str = "IINTS-AF Research Workbook") -> Path:
    """Apply academic workbook styling to an existing .xlsx file."""
    workbook_path = Path(path)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except Exception as exc:  # pragma: no cover - optional dependency
        raise RuntimeError("Excel styling requires openpyxl.") from exc

    wb = load_workbook(workbook_path)
    wb.properties.title = title
    wb.properties.creator = "IINTS-AF SDK"
    wb.properties.subject = "Pre-clinical research workbook"

    header_fill = PatternFill("solid", fgColor=_rgb_hex(ACADEMIC_THEME.navy))
    subheader_fill = PatternFill("solid", fgColor=_rgb_hex(ACADEMIC_THEME.soft_blue))
    alternate_fill = PatternFill("solid", fgColor=_rgb_hex(ACADEMIC_THEME.soft_gray))
    white_font = Font(color="FFFFFF", bold=True)
    text_font = Font(color=_rgb_hex(ACADEMIC_THEME.text))
    muted_font = Font(color=_rgb_hex(ACADEMIC_THEME.muted), italic=True)
    thin_rule = Side(style="thin", color=_rgb_hex(ACADEMIC_THEME.rule))
    border = Border(bottom=thin_rule)

    for ws in wb.worksheets:
        if ws.max_row == 0 or ws.max_column == 0:
            continue
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        ws.auto_filter.ref = ws.dimensions
        ws.row_dimensions[1].height = 24

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row_index = row[0].row
            for cell in row:
                if row_index % 2 == 0:
                    cell.fill = alternate_fill
                cell.font = text_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
                if isinstance(cell.value, float):
                    cell.number_format = "0.000"
                elif isinstance(cell.value, int):
                    cell.number_format = "0"

        for column_cells in ws.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_len = 10
            for cell in column_cells:
                if cell.value is not None:
                    max_len = max(max_len, min(60, len(str(cell.value))))
            ws.column_dimensions[column_letter].width = min(44, max_len + 2)

        ws.insert_rows(1)
        ws["A1"] = title
        ws["A1"].font = Font(color=_rgb_hex(ACADEMIC_THEME.navy), bold=True, size=13)
        ws["A1"].fill = subheader_fill
        ws["A1"].alignment = Alignment(vertical="center")
        if ws.max_column > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        ws.row_dimensions[1].height = 22
        ws["A2"].font = white_font
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"

        note_row = ws.max_row + 2
        ws.cell(note_row, 1).value = "Research-use only; not a medical-device output or treatment recommendation."
        ws.cell(note_row, 1).font = muted_font

    wb.save(workbook_path)
    return workbook_path
